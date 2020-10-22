from django.core.management.base import BaseCommand, CommandError
from prj.settings import DATA_DIR
from openpyxl import load_workbook
from market.models import Category, Product

class Command(BaseCommand):

    def handle(self, *args, **options):
        print('Clearing DB')
        Category.objects.all().delete()
        Product.objects.all().delete()

        print('Start importing from excel %s' % DATA_DIR)
        wb = load_workbook(DATA_DIR+'/init.xlsx')
        sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
        cat = None
        for i in range(1, sheet.max_row):
            item = sheet.cell(row=i, column=1).value
            price = sheet.cell(row=i, column=2).value
            if price == None:
                print('Create a new category')
                cat = Category()
                cat.name = item
                cat.save()
                cat = cat
            else:
                print('Create a new goods')
                if cat:
                    p = Product()
                    p.name = item
                    p.category = cat
                    p.save()



